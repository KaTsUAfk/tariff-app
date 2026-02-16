"""
Сервисы для работы с данными
"""
from typing import List, Dict, Optional
from datetime import datetime
import csv
import os

class RouteService:
    """Сервис для работы с маршрутами"""
    
    def __init__(self, db):
        self.db = db
    
    def get_all_routes(self) -> List[Dict]:
        """Получить все маршруты"""
        return self.db.get_all_routes()
    
    def get_route_by_id(self, route_id: int) -> Optional[Dict]:
        """Получить маршрут по ID"""
        return self.db.get_route_by_id(route_id)
    
    def add_route(self, number: str, name: str) -> int:
        """Добавить маршрут"""
        return self.db.add_route(number, name)
    
    def update_route(self, route_id: int, number: str, name: str) -> bool:
        """Обновить маршрут"""
        return self.db.update_route(route_id, number, name)
    
    def delete_route(self, route_id: int) -> bool:
        """Удалить маршрут"""
        return self.db.delete_route(route_id)
    
    def duplicate_route(self, source_route_id: int) -> int:
        """Дублировать маршрут со всеми пунктами"""
        source_route = self.get_route_by_id(source_route_id)
        if not source_route:
            raise ValueError("Исходный маршрут не найден")
        
        source_sequence = self.db.get_route_sequence(source_route_id)
        
        new_number = f"Копия {source_route['route_number']}"
        new_name = source_route['route_name']
        
        new_route_id = self.add_route(new_number, new_name)
        
        for point in source_sequence:
            self.db.add_point_to_route(
                new_route_id,
                point['point_id'],
                point['distance_km'],
                point['rounding'],
                point['cost_per_km'],
                point['baggage_percent']
            )
        
        return new_route_id


class PointService:
    """Сервис для работы с пунктами"""
    
    def __init__(self, db):
        self.db = db
    
    def get_all_points(self) -> List[Dict]:
        """Получить все пункты"""
        return self.db.get_all_points()
    
    def add_point(self, name: str) -> int:
        """Добавить пункт"""
        return self.db.add_point(name)
    
    def update_point(self, point_id: int, name: str) -> bool:
        """Обновить пункт"""
        return self.db.update_point(point_id, name)
    
    def delete_point(self, point_id: int) -> bool:
        """Удалить пункт"""
        return self.db.delete_point(point_id)
    
    def search_points(self, query: str) -> List[Dict]:
        """Поиск пунктов"""
        return self.db.search_points(query)
    
    def get_or_create_point(self, name: str) -> int:
        """Получить существующий пункт или создать новый"""
        points = self.search_points(name)
        for point in points:
            if point['name'].lower() == name.lower():
                return point['id']
        return self.add_point(name)


class ExportService:
    """Сервис для экспорта данных"""
    
    @staticmethod
    def export_to_csv(filename: str, data: List[Dict], headers: List[str]) -> None:
        """Экспорт в CSV"""
        with open(filename, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(headers)
            for row in data:
                writer.writerow([str(row.get(h, '')) for h in headers])
    
    @staticmethod
    def export_to_excel(filename: str, data: List[Dict], headers: List[str], 
                        sheet_name: str = "Данные") -> None:
        """Экспорт в Excel"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Стили
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Данные
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header, '')
                cell.border = thin_border
        
        # Автоподбор ширины
        for col in range(1, len(headers) + 1):
            max_length = 0
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2
        
        wb.save(filename)
    
    @staticmethod
    def generate_filename(prefix: str, suffix: str = "", extension: str = "csv") -> str:
        """Сгенерировать имя файла с датой"""
        date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        return f"{prefix}_{date_str}{('_' + suffix) if suffix else ''}.{extension}"


class ImportService:
    """Сервис для импорта данных"""
    
    @staticmethod
    def import_from_csv(filename: str, delimiter: str = ';') -> List[Dict]:
        """Импорт из CSV"""
        data = []
        with open(filename, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            for row in reader:
                data.append(row)
        return data
    
    @staticmethod
    def import_from_excel(filename: str) -> List[Dict]:
        """Импорт из Excel"""
        import openpyxl
        
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb.active
        
        headers = []
        data = []
        
        # Читаем заголовки
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers.append(str(header))
        
        # Читаем данные
        for row in range(2, ws.max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                value = ws.cell(row=row, column=col).value
                row_data[header] = str(value) if value is not None else ""
            if any(row_data.values()):
                data.append(row_data)
        
        return data
    
    @staticmethod
    def detect_delimiter(filename: str) -> str:
        """Определить разделитель в CSV файле"""
        with open(filename, 'r', encoding='utf-8-sig') as f:
            sample = f.read(1024)
            if ';' in sample:
                return ';'
            elif ',' in sample:
                return ','
            return ';'