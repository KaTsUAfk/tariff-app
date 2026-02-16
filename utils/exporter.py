"""
Экспорт таблицы стоимости проезда
"""
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from typing import List, Dict
import os

try:
    pdfmetrics.registerFont(TTFont('DejaVu', 'DejaVuSans.ttf'))
    PDF_FONT = 'DejaVu'
except:
    PDF_FONT = 'Helvetica'

class TariffExporter:
    @staticmethod
    def export_tariff_table(grid_info: Dict, points: List[Dict], 
                           matrix: Dict, tariffs_data: List, filename: str):
        """Экспорт таблицы стоимости в PDF (как в примере из документа)"""
        c = canvas.Canvas(filename, pagesize=landscape(A4))
        width, height = landscape(A4)
        
        # Заголовок
        c.setFont(PDF_FONT, 16)
        c.drawString(30, height - 40, f"Тарифная сетка №{grid_info['grid_number']} — {grid_info['grid_name']}")
        c.setFont(PDF_FONT, 10)
        c.drawString(30, height - 60, f"Пассажирский тариф: ₽{grid_info['passenger_tariff']:.2f} за 1 км")
        c.drawString(30, height - 75, f"Детская скидка: {grid_info['child_discount_percent']}% | Льготная скидка: {grid_info['benefit_discount_percent']}%")
        c.drawString(30, height - 90, f"Дата формирования: {datetime.now().strftime('%d.%m.%Y')}")
        
        # Подготовка данных таблицы
        n = len(points)
        table_data = [["Откуда \\ Куда"] + [p['name'] for p in points]]
        
        for i, row in enumerate(tariffs_data):
            table_row = [row[0]]  # Название пункта "откуда"
            for j in range(1, n + 1):
                cell = row[j]
                if cell['distance'] > 0:
                    text = f"{cell['base']:.2f}\n({cell['child']:.2f} / {cell['benefit']:.2f})"
                else:
                    text = "-"
                table_row.append(text)
            table_data.append(table_row)
        
        # Создание таблицы
        col_width = (width - 60) / (n + 1)
        col_widths = [col_width * 1.5] + [col_width] * n
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.2, 0.2)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), PDF_FONT),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, 1), (-1, -1), PDF_FONT),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ]))
        
        table.wrapOn(c, width - 60, height - 150)
        table.drawOn(c, 30, height - 450)
        
        # Футер
        c.setFont(PDF_FONT, 8)
        c.setFillColor(colors.grey)
        c.drawString(30, 30, "Таблица стоимости проезда между пунктами тарифной сетки")
        c.drawString(width - 200, 30, f"Страница 1 из 1")
        
        c.save()
        return filename
    
    @staticmethod
    def export_tariff_excel(grid_info: Dict, points: List[Dict],
                           matrix: Dict, tariffs_data: List, filename: str):
        """Экспорт в Excel с форматированием"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"Сетка {grid_info['grid_number']}"
        
        # Заголовок
        ws.merge_cells('A1:E1')
        ws['A1'] = f"ТАРИФНАЯ СЕТКА №{grid_info['grid_number']} — {grid_info['grid_name']}"
        ws['A1'].font = Font(name='Arial', size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.append([])
        ws.append([
            f"Пассажирский тариф: ₽{grid_info['passenger_tariff']:.2f}/км",
            f"Детская скидка: {grid_info['child_discount_percent']}%",
            f"Льготная скидка: {grid_info['benefit_discount_percent']}%",
            "",
            f"Дата: {datetime.now().strftime('%d.%m.%Y')}"
        ])
        ws.append([])
        
        # Таблица стоимости
        n = len(points)
        # Заголовки
        headers = ["Откуда \\ Куда"] + [p['name'] for p in points]
        ws.append(headers)
        
        header_fill = PatternFill(start_color="252628", end_color="252628", fill_type="solid")
        header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        
        for col in range(1, n + 2):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Данные
        for row_data in tariffs_data:
            ws.append(row_data)
        
        # Форматирование
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=n+1):
            for cell in row:
                cell.font = Font(name='Arial', size=9)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'), 
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Автоширина
        for col in range(1, n + 2):
            ws.column_dimensions[chr(64 + col)].width = 15 if col == 1 else 12
        
        wb.save(filename)
        return filename
    
    @staticmethod
    def get_suggested_filename(grid_number: str, ext: str = "pdf") -> str:
        timestamp = datetime.now().strftime("%Y%m%d")
        safe_number = "".join(c if c.isalnum() else "_" for c in grid_number)
        return f"tariff_grid_{safe_number}_{timestamp}.{ext}"